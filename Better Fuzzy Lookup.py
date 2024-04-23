# Better Fuzzy Lookup

from thefuzz import fuzz
from thefuzz import process
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys
import time

################################################
# Classes
################################################

class FasterFuzzy(tk.Tk):

    # Frame construction
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.title("Faster Fuzzy")
        self.geometry('500x400')
        self.frames = {}

        for F in ({MainPage}):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame(MainPage)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()    

class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        self.widget.configure(state="normal")
        self.widget.insert("end", str, (self.tag,))

    def flush(self):
        pass

class Workbook:
    """
    Class for representing an excel workbook within this script

    Parameters
    ----------
    path : str
        a string path to the file that will be represented by this class

    Constructor will create instances of the Tables class for each sheet present in workbook
    """
    matchedText = 'Rows Matched: 0'

    def __init__(self, path:str) -> None:
        self.path = path
        self.tables = {}
        self.__read()
        print('Workbook Loaded!')
    
    def __read(self):
        with pd.ExcelFile(self.path) as xls:
            for i, sheet in enumerate(xls.sheet_names):
                print(f'Reading sheet \'{sheet}\'')
                sheetdata = pd.read_excel(xls, sheet_name=sheet)
                self.tables[sheet] = Tables(sheet, i)
                self.tables[sheet].readData(sheetdata)

    def getSheets(self) -> list:
        return list(self.tables.keys())

    def getPath(self) -> str:
        return self.path
    
    def setMatchedCount(self, count) -> str:
        self.matchedText = f'Rows Matched: {count}'

    def getMatchedCount(self) -> str:
        return self.matchedText
    

class Tables:
    """
    Tables class that represents a workbook's sheets and the data stored in them

    Parameters
    ----------
    table_name : str
        A string name for this instance table
    table_id : int
        An interger id to index tables
    """
    def __init__(self, table_name, table_id) -> None:
        self.table_name = table_name
        self.table_id = table_id
        self.data = pd.DataFrame()

    def __readColumnHeader(self):
        print(f"Extracting columns from \'{self.table_name}\'...\n")
        oldColList = self.data.columns.values.tolist()
        newColList = [f"{col}.{self.table_name}" for col in oldColList]
        oldNewDict = {}
        for old, new in zip(oldColList, newColList):
            oldNewDict[old] = new
        self.data.rename(columns=oldNewDict, inplace=True)
        self.columns = self.data.columns.values.tolist()

    def readData(self, sheetData: pd.DataFrame):
        print("Reading table data...")
        self.data = sheetData
        self.__readColumnHeader()
    
    def getHeaders(self) -> list:
        return self.columns
    
    def getData(self) -> pd.DataFrame:
        return self.data
    
    def getName(self) -> str:
        return self.table_name

################################################
# Helper functions
################################################

def chooseFileHandler(filePathLabel:tk.Label) -> str:
    """
    Prompts user to select a file to load

    Parameters
    ----------
    filePathLabel : tk.Label
        the GUI label that should be updated to display the filepath
    -------
    Returns
    file path string
    """
    excelFile = askopenfilename()
    if(excelFile == ""):
        filePathLabel['text'] = ""
    elif(".xlsx" not in excelFile):
        tk.messagebox.showerror("Error", "Incorrect File Format")
        filePathLabel['text'] = ""
        excelFile = ""
    return excelFile

def writeData(data:pd.DataFrame, excelFilePath:str, sheet_name:str) -> None:
    """
    Writes dataframe to an excel sheet of file

    Parameters
    ----------
    data : pandas dataframe
        dataframe to write to file
    excelFilePath : str
        path to excel file
    sheet_name : str
        name of sheet that will be created and wrote to
    """
    print("Updating Data...\n")
    options = {}
    options['strings_to_formulas'] = False
    options['strings_to_urls'] = False
    with pd.ExcelWriter(excelFilePath,
                        mode='a',
                        engine='openpyxl',
                        if_sheet_exists='new',
    ) as writer:
        writer.workbook = load_workbook(excelFilePath)
        data.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        print("Data Updated.\nClosing Workbook...\n") 
    print("Work Book Closed.\n")

def runButtonHandler(Workbook: Workbook, sheet_selection:list, column_selection:list, similarity_threshold:int, matchLimit:int) -> None:
    """
    Handler function to execute main program logic

    Paramerters
    -----------
    Workbook : Worbook
        A workbook object instance from this file's workbook class
    sheet_selection : list
        list of strings, length 2
    column_selection : list
        list of strings, length 2
    similarity_threshold : int
        between 0 and 100
    matchLimit : int
        number of matches to find
    """
    start_time = time.perf_counter()
    tableData = []
    names = []
    for sheet in sheet_selection:
        tableData.append(Workbook.tables[sheet].getData())
        names.append(Workbook.tables[sheet].getName())
    
    matchedtable = fuzzyMatch(tableData, column_selection, similarity_threshold, matchLimit, Workbook)
    writeData(matchedtable, Workbook.getPath(), f"FLU_{names[0]}_{names[1]}")
    print("execution time = %s seconds" % (time.perf_counter() - start_time))

#original
def fuzzyMatch(tabledata:list, column_selection:list, similarity_threshold:int, matchLimit:int, wb:Workbook) -> pd.DataFrame:
    """
    Preforms the fuzzy match between two tables
    """
    left_table = tabledata[0].fillna("")
    right_table = tabledata[1].fillna("")
    matchOnLeft = column_selection[0]
    matchOnRight = column_selection[1]

    matchedCounter = 0

    combinedTable = pd.DataFrame(columns=[*left_table.columns.values, *right_table.columns.values, 'Similarity Score'])
    
    for index, val in left_table.iterrows():
        returnedValues = process.extract(val[matchOnLeft], right_table[matchOnRight], scorer=fuzz.token_sort_ratio, limit=matchLimit)
        filteredValues = []
        skippedPairs = 0
        for pair in returnedValues:
            if pair[1] >= similarity_threshold or len(returnedValues) - skippedPairs == 1:
                filteredValues.append(pair)
            else:
                skippedPairs += 1

        for pair in filteredValues:
            if len(filteredValues) == 1 and pair[1] < similarity_threshold:
                pair = ("", 0)
                combinedrow = left_table.iloc[index]
                combinedrow['Similarity Score'] = pair[1]
                combinedTable = combinedTable.append(combinedrow, ignore_index=True)
            else:
                row_from_left = left_table.iloc[index]
                row_from_right = right_table.iloc[pair[2]]
                combinedrow = pd.concat([row_from_left, row_from_right])
                combinedrow['Similarity Score'] = pair[1]
                combinedTable = combinedTable.append(combinedrow, ignore_index=True)
                matchedCounter += 1
    wb.setMatchedCount(matchedCounter)
    return combinedTable


################################################
# GUI Window
################################################

class MainPage(tk.Frame):

    #backgroundcolour = "#fcba03"
    backgroundcolour = "light gray"

    def initValues(self):
        self.similarity_threshold = 50
        self.num_matches = 1
        self.sheetOptions = []
        self.colOptions_1 = []
        self.colOptions_2 = []
        self.excelFilePath = ""
        self.matchLimit = 1
        self.selected_tables = ["",""]
        self.selected_columns = ["",""]
        self.selectedsheet1 = tk.StringVar()
        self.selectedsheet2 = tk.StringVar()
        self.selectedcol1 = tk.StringVar()
        self.selectedcol2 = tk.StringVar()

    #def similaritySlider(self):
    
    def onRunPress(self):
        self.similarity_threshold = int(self.similarity_slider.get())
        self.matchLimit = int(self.limit_spinbox.get())
        self.selected_columns[0] = self.colselectorbox_1.get()
        self.selected_columns[1] = self.colselectorbox_2.get()
        self.selected_tables[0] = self.sheetselectorbox_1.get()
        self.selected_tables[1] = self.sheetselectorbox_2.get()

        runButtonHandler(self.thisWorkbook, self.selected_tables, self.selected_columns, self.similarity_threshold, self.matchLimit)
        self.matchedCountLabel['text'] = self.thisWorkbook.getMatchedCount()

    def onChooseFilePress(self):
        #sys.stdout = TextRedirector()
        self.excelFilePath = chooseFileHandler(tk.Label(self))
        self.sourcefilelabeldisp['text'] = self.excelFilePath
        print(f'Reading Workbook \'{self.excelFilePath}\'')

        self.thisWorkbook = Workbook(self.excelFilePath)
        self.sheetOptions = self.thisWorkbook.getSheets()

        self.sheetselectorbox_1['values'] = self.sheetOptions
        self.sheetselectorbox_2['values'] = self.sheetOptions
        self.selectedsheet1 = self.sheetOptions[0]
        self.selectedsheet2 = self.sheetOptions[1]
        self.sheetselectorbox_1.set(self.selectedsheet1)
        self.sheetselectorbox_2.set(self.selectedsheet2)

        self.colselectorbox_1.set(self.thisWorkbook.tables[self.selectedsheet1].getHeaders()[0])
        self.colselectorbox_2.set(self.thisWorkbook.tables[self.selectedsheet2].getHeaders()[0])

        self.sheetselectorbox_1.bind('<<ComboboxSelected>>', self.onSheetSelect_1)
        self.sheetselectorbox_2.bind('<<ComboboxSelected>>', self.onSheetSelect_2)
        self.checkToEnableRun()

    def onSheetSelect_1(self, event):
        if self.excelFilePath == '': return
        self.sheetselectorbox_1_val = event.widget.get()
        #self.colOptions_1 = readColumns(self.excelFile, self.sheetselectorbox_1_val)
        self.colOptions_1 = self.thisWorkbook.tables[self.sheetselectorbox_1_val].getHeaders()
        self.colselectorbox_1['values'] = self.colOptions_1
        self.selectedcol1 = self.colOptions_1[0]
        self.colselectorbox_1.set(self.selectedcol1)
        self.checkToEnableRun()

    def onSheetSelect_2(self, event):
        if self.excelFilePath == '': return
        self.sheetselectorbox_2_val = event.widget.get()
        #self.colOptions_2 = readColumns(self.excelFile, self.sheetselectorbox_2_val)
        self.colOptions_2 = self.thisWorkbook.tables[self.sheetselectorbox_2_val].getHeaders()
        self.colselectorbox_2['values'] = self.colOptions_2
        self.selectedcol2 = self.colOptions_2[0]
        self.colselectorbox_2.set(self.selectedcol2)
        self.checkToEnableRun()

    def checkToEnableRun(self, event=None):
        if self.excelFilePath != '':
            if self.sheetselectorbox_1.get() != self.sheetselectorbox_2.get():
                self.runButton['state'] = 'normal'
            else:
                self.runButton['state'] = 'disabled'

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.initValues()

        ################################
        # Row 0
        ################################
        header = tk.Label(self, text="Faster Fuzzy", font="Arial 20 bold")
        header.configure(background=self.backgroundcolour)
        header.grid(column=1, row=0, pady=10)

        ################################
        # Row 1
        ################################

        chooseFileButton = tk.Button(self, text="Choose File", font="Arial 14", command=self.onChooseFilePress)
        chooseFileButton.grid(row=1, column=0)

        self.runButton = tk.Button(self, text="Run Fuzzy Lookup", font="Arial 14", command=self.onRunPress)
        self.runButton.grid(row=1, column=2)
        self.runButton['state'] = 'disabled'

        ################################
        # Row 2
        ################################
        
        self.sourcefilelabel = tk.Label(self)
        self.sourcefilelabel.configure(background=self.backgroundcolour)
        self.sourcefilelabel.grid(column=0, row=2,sticky="nw")
        self.sourcefilelabel['text'] = 'File Path:'

        self.sourcefilelabeldisp = tk.Label(self)
        self.sourcefilelabeldisp.configure(background=self.backgroundcolour)
        self.sourcefilelabeldisp.grid(column=1, row=2,sticky="nw")

        ################################
        # Row 3
        ################################

        self.similarity_slider = tk.Scale(self, from_=0, to=100, orient="horizontal")
        self.similarity_slider.grid(row = 5, column=1)
        self.similarity_slider.set(self.similarity_threshold)

        self.limit_spinbox = tk.Spinbox(self, from_=1, to=5)
        self.limit_spinbox.grid(row=7, column=1)

        ################################
        # Row 4
        ################################
        
        #combo boxes
        self.sheetselectorbox_1 = ttk.Combobox(self, textvariable=self.selectedsheet1)
        self.sheetselectorbox_1['values'] = self.sheetOptions
        self.sheetselectorbox_1.state(["readonly"])
        self.sheetselectorbox_1.grid(row=5, column=0)

        self.sheetselectorbox_2 = ttk.Combobox(self, textvariable=self.selectedsheet2)
        self.sheetselectorbox_2['values'] = self.sheetOptions
        self.sheetselectorbox_2.state(["readonly"])
        self.sheetselectorbox_2.grid(row=5, column=2)

        self.tableLLabel = tk.Label(self)
        self.tableLLabel.configure(background=self.backgroundcolour)
        self.tableLLabel.grid(column=0, row=4,sticky="nw")
        self.tableLLabel['text'] = 'Left Table'

        self.tableRLabel = tk.Label(self)
        self.tableRLabel.configure(background=self.backgroundcolour)
        self.tableRLabel.grid(column=2, row=4,sticky="nw")
        self.tableRLabel['text'] = 'Right Table'

        self.colLLabel = tk.Label(self)
        self.colLLabel.configure(background=self.backgroundcolour)
        self.colLLabel.grid(column=0, row=6,sticky="nw")
        self.colLLabel['text'] = 'Left Matching Column'

        self.colRLabel = tk.Label(self)
        self.colRLabel.configure(background=self.backgroundcolour)
        self.colRLabel.grid(column=2, row=6,sticky="nw")
        self.colRLabel['text'] = 'Right Matching Column'

        self.colselectorbox_1 = ttk.Combobox(self, textvariable=self.selectedcol1)
        self.colselectorbox_1['values'] = self.colOptions_1
        self.colselectorbox_1.state(["readonly"])
        self.colselectorbox_1.bind('<<ComboboxSelected>>', self.checkToEnableRun)
        self.colselectorbox_1.grid(row=7, column=0)

        self.colselectorbox_2 = ttk.Combobox(self, textvariable=self.selectedcol2)
        self.colselectorbox_2['values'] = self.colOptions_2
        self.colselectorbox_2.state(["readonly"])
        self.colselectorbox_2.bind('<<ComboboxSelected>>', self.checkToEnableRun)
        self.colselectorbox_2.grid(row=7, column=2)

        self.sliderLabel = tk.Label(self)
        self.sliderLabel.configure(background=self.backgroundcolour)
        self.sliderLabel.grid(column=1, row=4,sticky="nw")
        self.sliderLabel['text'] = 'Similarity Cutoff'

        self.sliderLabel = tk.Label(self)
        self.sliderLabel.configure(background=self.backgroundcolour)
        self.sliderLabel.grid(column=1, row=6,sticky="nw")
        self.sliderLabel['text'] = 'Maximum Matches to return'

        self.matchedCountLabel = tk.Label(self)
        self.matchedCountLabel.configure(background=self.backgroundcolour)
        self.matchedCountLabel.grid(column=0, row=9,sticky="nw")
        self.matchedCountLabel['text'] = Workbook.matchedText
        
        self.configure(background=self.backgroundcolour)

        self.grid_columnconfigure(0, pad=10)
        self.grid_columnconfigure(1, weight=3)



if __name__ == '__main__':
    app = FasterFuzzy()
    app.mainloop()    